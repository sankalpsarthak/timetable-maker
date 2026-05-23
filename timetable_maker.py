import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar

class TimeTableMaker:
    """A comprehensive timetable and planner maker using Excel"""
    
    def __init__(self, filename="timetable.xlsx"):
        self.filename = filename
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        self.colors = {
            'header': 'FF4472C4',      # Blue
            'primary': 'FFC6D9F1',     # Light Blue
            'secondary': 'FFE7E6E6',   # Light Gray
            'accent': 'FFFFF2CC',      # Light Yellow
            'success': 'FFC6EFCE',     # Light Green
            'warning': 'FFFFEB9C'      # Light Orange
        }
        self.borders = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _apply_header_style(self, cell, bold=True):
        """Apply header styling to a cell"""
        cell.font = Font(bold=bold, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = self.borders
    
    def _apply_cell_style(self, cell, color='primary', alignment='center'):
        """Apply standard cell styling"""
        cell.fill = PatternFill(start_color=self.colors[color], end_color=self.colors[color], fill_type="solid")
        cell.alignment = Alignment(horizontal=alignment, vertical="center", wrap_text=True)
        cell.border = self.borders
    
    def create_daily_timetable(self, activities=None, time_interval=30):
        """Create a daily timetable sheet
        
        Args:
            activities (dict): Dictionary with time as key and activity as value
            time_interval (int): Time interval in minutes (15, 30, 60)
        """
        ws = self.workbook.create_sheet("Daily Timetable")
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        
        # Header
        ws['A1'] = "Daily Timetable"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        date_today = datetime.now().strftime("%A, %d %B %Y")
        ws['A2'] = f"Date: {date_today}"
        ws.merge_cells('A2:B2')
        
        # Column headers
        ws['A3'] = "Time"
        ws['B3'] = "Activity"
        self._apply_header_style(ws['A3'])
        self._apply_header_style(ws['B3'])
        
        # Generate time slots
        start_hour = 6
        end_hour = 23
        row = 4
        
        for hour in range(start_hour, end_hour + 1):
            for minute in range(0, 60, time_interval):
                time_str = f"{hour:02d}:{minute:02d}"
                ws[f'A{row}'] = time_str
                ws[f'B{row}'] = activities.get(time_str, "") if activities else ""
                
                self._apply_cell_style(ws[f'A{row}'], 'secondary', 'center')
                self._apply_cell_style(ws[f'B{row}'], 'primary', 'left')
                row += 1
    
    def create_weekly_timetable(self, activities=None):
        """Create a weekly timetable sheet
        
        Args:
            activities (dict): Dictionary with day as key and activities dict as value
        """
        ws = self.workbook.create_sheet("Weekly Timetable")
        
        # Header
        ws['A1'] = "Weekly Timetable"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        current_date = datetime.now()
        week_start = current_date - timedelta(days=current_date.weekday())
        week_str = f"{week_start.strftime('%d %B')} - {(week_start + timedelta(days=6)).strftime('%d %B %Y')}"
        ws['A2'] = f"Week: {week_str}"
        ws.merge_cells('A2:H2')
        
        # Days of week
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        ws['A3'] = "Time"
        self._apply_header_style(ws['A3'])
        
        col = 2
        for day in days:
            cell = ws.cell(row=3, column=col)
            cell.value = day
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col)].width = 18
            col += 1
        
        # Time slots
        start_hour = 6
        end_hour = 22
        row = 4
        
        for hour in range(start_hour, end_hour + 1):
            time_str = f"{hour:02d}:00"
            ws[f'A{row}'] = time_str
            self._apply_cell_style(ws[f'A{row}'], 'secondary', 'center')
            
            for col in range(2, 9):
                cell = ws.cell(row=row, column=col)
                cell.value = ""
                self._apply_cell_style(cell, 'primary', 'center')
            row += 1
    
    def create_planner(self, period="monthly"):
        """Create a planner sheet for different time periods
        
        Args:
            period (str): 'daily', 'weekly', 'monthly', '3monthly', '6monthly', '9monthly', '12monthly'
        """
        periods = {
            'daily': ('Daily Planner', 1, self._create_daily_planner),
            'weekly': ('Weekly Planner', 7, self._create_weekly_planner),
            'monthly': ('Monthly Planner', 30, self._create_monthly_planner),
            '3monthly': ('3-Monthly Planner', 90, self._create_multi_month_planner),
            '6monthly': ('6-Monthly Planner', 180, self._create_multi_month_planner),
            '9monthly': ('9-Monthly Planner', 270, self._create_multi_month_planner),
            '12monthly': ('Yearly Planner', 365, self._create_yearly_planner)
        }
        
        if period not in periods:
            raise ValueError(f"Period must be one of {list(periods.keys())}")
        
        sheet_name, duration, func = periods[period]
        ws = self.workbook.create_sheet(sheet_name)
        func(ws, duration)
    
    def _create_daily_planner(self, ws, duration):
        """Create daily planner layout"""
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        
        ws['A1'] = "Daily Planner"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        date_today = datetime.now().strftime("%A, %d %B %Y")
        ws['A2'] = f"Date: {date_today}"
        ws.merge_cells('A2:B2')
        
        sections = [
            "Morning Tasks",
            "Afternoon Tasks",
            "Evening Tasks",
            "Priority Items",
            "Notes"
        ]
        
        row = 4
        for section in sections:
            ws[f'A{row}'] = section
            self._apply_header_style(ws[f'A{row}'])
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            for i in range(3):
                ws[f'A{row}'] = "☐"
                ws[f'B{row}'] = ""
                self._apply_cell_style(ws[f'A{row}'], 'accent', 'center')
                self._apply_cell_style(ws[f'B{row}'], 'primary', 'left')
                ws.row_dimensions[row].height = 25
                row += 1
            row += 1
    
    def _create_weekly_planner(self, ws, duration):
        """Create weekly planner layout"""
        ws.column_dimensions['A'].width = 20
        for col in range(2, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        ws['A1'] = "Weekly Planner"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        # Days header
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        ws['A2'] = "Focus Areas"
        self._apply_header_style(ws['A2'])
        
        for col, day in enumerate(days, 2):
            cell = ws.cell(row=2, column=col)
            cell.value = day
            self._apply_header_style(cell)
        
        # Weekly sections
        sections = [
            "Work/Study",
            "Health/Fitness",
            "Personal Growth",
            "Relationships",
            "Leisure"
        ]
        
        row = 3
        for section in sections:
            ws[f'A{row}'] = section
            self._apply_header_style(ws[f'A{row}'])
            for col in range(2, 9):
                cell = ws.cell(row=row, column=col)
                cell.value = ""
                self._apply_cell_style(cell, 'primary', 'center')
                ws.row_dimensions[row].height = 30
            row += 1
    
    def _create_monthly_planner(self, ws, duration):
        """Create monthly planner layout"""
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        
        ws['A1'] = "Monthly Planner"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        now = datetime.now()
        month_year = now.strftime("%B %Y")
        ws['A2'] = f"Month: {month_year}"
        ws.merge_cells('A2:B2')
        
        # Days of month
        num_days = calendar.monthrange(now.year, now.month)[1]
        
        row = 4
        for day in range(1, num_days + 1):
            date_str = f"{day:02d} {now.strftime('%B')}"
            ws[f'A{row}'] = date_str
            ws[f'B{row}'] = ""
            self._apply_cell_style(ws[f'A{row}'], 'secondary', 'center')
            self._apply_cell_style(ws[f'B{row}'], 'primary', 'left')
            ws.row_dimensions[row].height = 25
            row += 1
    
    def _create_multi_month_planner(self, ws, duration):
        """Create multi-month planner layout"""
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        
        months = duration // 30
        ws['A1'] = f"{months}-Month Planner"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        now = datetime.now()
        start_date = now.strftime("%d %B %Y")
        end_date = (now + timedelta(days=duration)).strftime("%d %B %Y")
        ws['A2'] = f"Period: {start_date} to {end_date}"
        ws.merge_cells('A2:B2')
        
        # Monthly milestones
        row = 4
        current_date = now
        for month_num in range(months):
            month_str = current_date.strftime("%B %Y")
            ws[f'A{row}'] = month_str
            self._apply_header_style(ws[f'A{row}'])
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            for i in range(4):
                ws[f'A{row}'] = f"  Week {i+1}"
                ws[f'B{row}'] = ""
                self._apply_cell_style(ws[f'A{row}'], 'accent', 'left')
                self._apply_cell_style(ws[f'B{row}'], 'primary', 'left')
                ws.row_dimensions[row].height = 25
                row += 1
            
            row += 1
            current_date = current_date + timedelta(days=30)
    
    def _create_yearly_planner(self, ws, duration):
        """Create yearly planner layout"""
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        
        ws['A1'] = "Yearly Planner"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        now = datetime.now()
        year = now.year
        ws['A2'] = f"Year: {year}"
        ws.merge_cells('A2:B2')
        
        # Quarterly breakdown
        row = 4
        for quarter in range(1, 5):
            ws[f'A{row}'] = f"Q{quarter} ({['Jan-Mar', 'Apr-Jun', 'Jul-Sep', 'Oct-Dec'][quarter-1]})"
            self._apply_header_style(ws[f'A{row}'])
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            for i in range(3):
                ws[f'A{row}'] = f"  Goal {i+1}"
                ws[f'B{row}'] = ""
                self._apply_cell_style(ws[f'A{row}'], 'accent', 'left')
                self._apply_cell_style(ws[f'B{row}'], 'primary', 'left')
                ws.row_dimensions[row].height = 25
                row += 1
            
            row += 1
    
    def create_habit_tracker(self):
        """Create a habit tracker for the current month"""
        ws = self.workbook.create_sheet("Habit Tracker")
        
        now = datetime.now()
        month_year = now.strftime("%B %Y")
        num_days = calendar.monthrange(now.year, now.month)[1]
        
        ws['A1'] = f"Habit Tracker - {month_year}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A1:AM1')  # Adjust for all days
        
        ws.column_dimensions['A'].width = 20
        
        # Header with day numbers
        ws['A2'] = "Habit"
        self._apply_header_style(ws['A2'])
        
        for day in range(1, num_days + 1):
            col_letter = get_column_letter(day + 1)
            cell = ws[f'{col_letter}2']
            cell.value = day
            self._apply_header_style(cell)
            ws.column_dimensions[col_letter].width = 5
        
        # Add some default habits
        habits = [
            "Exercise",
            "Reading",
            "Meditation",
            "Healthy Eating",
            "Sleep 8 Hours",
            "Work Goals",
            "Family Time",
            "Hydration"
        ]
        
        row = 3
        for habit in habits:
            ws[f'A{row}'] = habit
            self._apply_cell_style(ws[f'A{row}'], 'secondary', 'left')
            
            for day in range(1, num_days + 1):
                col_letter = get_column_letter(day + 1)
                cell = ws[f'{col_letter}{row}']
                cell.value = "☐"  # Checkbox symbol
                self._apply_cell_style(cell, 'success', 'center')
                ws.row_dimensions[row].height = 20
            
            row += 1
        
        # Add empty rows for custom habits
        for i in range(3):
            ws[f'A{row}'] = ""
            self._apply_cell_style(ws[f'A{row}'], 'secondary', 'left')
            
            for day in range(1, num_days + 1):
                col_letter = get_column_letter(day + 1)
                cell = ws[f'{col_letter}{row}']
                cell.value = "☐"
                self._apply_cell_style(cell, 'success', 'center')
            
            row += 1
    
    def save(self):
        """Save the workbook"""
        self.workbook.save(self.filename)
        print(f"✓ Timetable saved to {self.filename}")
    
    def save_as(self, filename):
        """Save the workbook with a different name"""
        self.filename = filename
        self.save()